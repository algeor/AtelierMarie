"""Accounting export package builder tests."""

from pathlib import Path
import sqlite3

from openpyxl import load_workbook
import pytest


def _seed_reviewed_settings(db: sqlite3.Connection) -> tuple[int, int]:
    db.execute(
        """
        INSERT INTO seller_legal_profile_versions (
            effective_date, reviewed, legal_name, default_currency
        ) VALUES ('2026-08-01', 1, 'Atelier Marie OOD', 'EUR')
        """
    )
    seller_id = int(db.execute("SELECT last_insert_rowid()").fetchone()[0])
    db.execute(
        """
        INSERT INTO vat_fiscal_settings_versions (
            effective_date, reviewed, vat_mode, fiscal_document_mode
        ) VALUES ('2026-08-01', 1, 'registered', 'external_reference')
        """
    )
    vat_id = int(db.execute("SELECT last_insert_rowid()").fetchone()[0])
    db.commit()
    return seller_id, vat_id


def _seed_paid_order(db: sqlite3.Connection, app, *, seller_id: int, vat_id: int) -> None:
    db.execute(
        "INSERT OR IGNORE INTO products (id, name_en, price_cents, stock) VALUES ('export-candle', 'Export Candle', 1000, 10)"
    )
    db.execute(
        """
        INSERT INTO orders (
            id, session_id, status, total_cents, customer_email, customer_name,
            payment_method, payment_status, seller_legal_profile_version_id,
            vat_fiscal_settings_version_id, accounting_classification_state,
            accounting_readiness_status, created_at, updated_at
        ) VALUES ('export-order', ?, 'confirmed', 1000, 'export@example.com',
                  'Export Buyer', 'card', 'paid', ?, ?, 'domestic_default', 'ready',
                  '2026-08-10 10:00:00', '2026-08-10 10:00:00')
        """,
        (app._test_session_id, seller_id, vat_id),
    )
    db.execute(
        """
        INSERT INTO order_items (order_id, product_id, product_name, price_cents, quantity)
        VALUES ('export-order', 'export-candle', 'Export Candle', 1000, 1)
        """
    )
    db.execute(
        """
        INSERT INTO payments (id, order_id, provider, amount_cents, provider_status)
        VALUES ('export-payment', 'export-order', 'stripe', 1000, 'paid')
        """
    )
    db.execute(
        """
        INSERT INTO expense_evidence (
            id, supplier_name, document_number, purchase_date, payment_status,
            category_key, gross_amount_cents, tax_amount_cents, review_status
        ) VALUES ('export-expense', 'Wax Supplier', 'SUP-EXP', '2026-08-05',
                  'paid', 'materials', 12000, 2000, 'reviewed')
        """
    )
    db.commit()


def _seed_inventory_export_data(db: sqlite3.Connection) -> None:
    db.execute(
        """
        UPDATE inventory_settings
        SET valuation_enabled = 1, accountant_reviewed = 1, valuation_method = 'weighted_average'
        WHERE id = 'default'
        """
    )
    db.execute(
        """
        INSERT INTO product_inventory_profiles (
            product_id, inventory_mode, stock_source, opening_balance_state, valuation_readiness
        ) VALUES ('export-candle', 'ledger_managed', 'inventory_ledger', 'reviewed', 'ready')
        """
    )
    db.execute(
        """
        INSERT INTO materials (id, sku, name, category, stock_uom)
        VALUES ('export-wax', 'EXP-WAX', 'Export Wax', 'wax', 'g')
        """
    )
    db.execute(
        """
        INSERT INTO inventory_movements (
            id, item_type, item_id, movement_type, quantity_delta, uom,
            source_type, source_id, review_state, occurred_at
        ) VALUES ('export-wax-receipt-move', 'material', 'export-wax', 'receipt',
                  1000, 'g', 'material_receipt', 'export-wax-receipt', 'reviewed',
                  '2026-08-03 09:00:00')
        """
    )
    db.execute(
        """
        INSERT INTO inventory_movements (
            id, item_type, item_id, movement_type, quantity_delta, uom,
            source_type, source_id, review_state, occurred_at
        ) VALUES ('export-wax-writeoff-move', 'material', 'export-wax', 'write_off',
                  -100, 'g', 'stock_count', 'export-count', 'reviewed',
                  '2026-08-20 09:00:00')
        """
    )
    db.execute(
        """
        INSERT INTO inventory_movements (
            id, item_type, item_id, movement_type, quantity_delta, uom,
            product_id, order_id, order_item_key, review_state, occurred_at
        ) VALUES ('export-sale-move', 'finished_good', 'export-candle', 'sale_issue',
                  -1, 'unit', 'export-candle', 'export-order', 'export-order:export-candle',
                  'reviewed', '2026-08-10 10:00:00')
        """
    )
    db.execute(
        """
        INSERT INTO inventory_valuation_layers (
            id, movement_id, item_type, item_id, quantity, unit_value_amount,
            total_value_cents, currency, valuation_method, source_type, source_id,
            valuation_date, review_state
        ) VALUES ('export-wax-receipt-layer', 'export-wax-receipt-move', 'material',
                  'export-wax', 1000, '0.010000', 1000, 'EUR', 'weighted_average',
                  'material_receipt', 'export-wax-receipt', '2026-08-03', 'official')
        """
    )
    db.execute(
        """
        INSERT INTO inventory_valuation_layers (
            id, movement_id, item_type, item_id, quantity, unit_value_amount,
            total_value_cents, currency, valuation_method, source_type, source_id,
            valuation_date, review_state
        ) VALUES ('export-wax-writeoff-layer', 'export-wax-writeoff-move', 'material',
                  'export-wax', -100, '0.010000', 100, 'EUR', 'weighted_average',
                  'stock_count', 'export-count', '2026-08-20', 'official')
        """
    )
    db.execute(
        """
        INSERT INTO inventory_valuation_layers (
            id, movement_id, item_type, item_id, quantity, unit_value_amount,
            total_value_cents, currency, valuation_method, source_type, source_id,
            valuation_date, review_state
        ) VALUES ('export-finished-opening-layer', NULL, 'finished_good', 'export-candle',
                  5, '2.000000', 1000, 'EUR', 'weighted_average', 'opening_balance',
                  'export-opening', '2026-08-01', 'official')
        """
    )
    db.execute(
        """
        INSERT INTO inventory_valuation_layers (
            id, movement_id, item_type, item_id, quantity, unit_value_amount,
            total_value_cents, currency, valuation_method, source_type, source_id,
            valuation_date, review_state
        ) VALUES ('export-sale-layer', 'export-sale-move', 'finished_good', 'export-candle',
                  -1, '2.000000', 200, 'EUR', 'weighted_average', 'order_item',
                  'export-order:export-candle', '2026-08-10', 'official')
        """
    )
    db.execute(
        """
        INSERT INTO cogs_ledger (
            id, order_id, order_number, order_item_key, product_id, quantity_sold,
            cogs_date, unit_cost_amount, total_cost_cents, currency, valuation_method,
            source_movement_id, source_valuation_layer_id, review_state
        ) VALUES ('export-cogs', 'export-order', 'EXPORT-ORDER', 'export-order:export-candle',
                  'export-candle', 1, '2026-08-10', '2.000000', 200, 'EUR',
                  'weighted_average', 'export-sale-move', 'export-sale-layer', 'official')
        """
    )
    db.commit()


async def _closed_period(admin_client) -> str:
    period_resp = await admin_client.post(
        "/v1/admin/accounting/periods",
        json={"period_start": "2026-08-01", "period_end": "2026-08-31", "currency": "EUR"},
    )
    period_id = period_resp.json()["id"]
    await admin_client.post(f"/v1/admin/accounting/periods/{period_id}/review")
    close_resp = await admin_client.post(f"/v1/admin/accounting/periods/{period_id}/close")
    assert close_resp.status_code == 200
    return period_id


@pytest.mark.asyncio
async def test_export_package_generation_manifest_download_accept_and_versioning(admin_client, db, app):
    seller_id, vat_id = _seed_reviewed_settings(db)
    _seed_paid_order(db, app, seller_id=seller_id, vat_id=vat_id)
    _seed_inventory_export_data(db)

    open_period_resp = await admin_client.post(
        "/v1/admin/accounting/periods",
        json={"period_start": "2026-07-01", "period_end": "2026-07-31", "currency": "EUR"},
    )
    open_export_resp = await admin_client.post(
        f"/v1/admin/accounting/periods/{open_period_resp.json()['id']}/exports"
    )
    assert open_export_resp.status_code == 409
    assert open_export_resp.json()["error"]["code"] == "PERIOD_MUST_BE_CLOSED"

    period_id = await _closed_period(admin_client)
    export_resp = await admin_client.post(f"/v1/admin/accounting/periods/{period_id}/exports")
    assert export_resp.status_code == 200
    package = export_resp.json()
    assert package["version"] == 1
    assert package["current_final"] is True
    xlsx_path = Path(package["xlsx_path"])
    manifest_path = Path(package["manifest_path"])
    assert xlsx_path.exists()
    assert manifest_path.exists()
    assert "private-exports" in str(xlsx_path)
    manifest = package["manifest"]
    assert manifest["schema_version"] == "accounting-finance-hub.v1"
    assert manifest["row_counts"]["sales"] >= 1
    assert manifest["row_counts"]["inventory_movements"] == 3
    assert manifest["row_counts"]["material_on_hand"] == 1
    assert manifest["row_counts"]["finished_goods_on_hand"] == 1
    assert manifest["row_counts"]["inventory_valuation"] == 4
    assert manifest["row_counts"]["cogs"] == 1
    assert manifest["row_counts"]["inventory_writeoffs"] == 1
    assert manifest["sheet_totals"]["cogs"]["total_cost_cents"] == 200
    assert manifest["files"]["csv_components"]["inventory_movements.csv"]["sha256"]
    assert manifest["files"]["csv_components"]["cogs.csv"]["totals"]["total_cost_cents"] == 200
    assert "summary.csv" in manifest["files"]["csv_components"]
    assert manifest["files"]["xlsx"]["sha256"]

    workbook = load_workbook(xlsx_path, read_only=True)
    assert {
        "summary",
        "sales",
        "expenses",
        "product_costs",
        "settings_snapshot",
        "inventory_movements",
        "material_on_hand",
        "finished_goods_on_hand",
        "inventory_valuation",
        "cogs",
        "inventory_writeoffs",
    }.issubset(set(workbook.sheetnames))
    workbook.close()

    download_resp = await admin_client.get(
        f"/v1/admin/accounting/exports/{package['id']}/download?file=manifest"
    )
    assert download_resp.status_code == 200
    assert download_resp.headers["cache-control"] == "no-store, no-cache"

    accept_resp = await admin_client.post(
        f"/v1/admin/accounting/exports/{package['id']}/accept",
        json={
            "accountant_name": "Accountant Demo",
            "accountant_reference": "AUG-OK",
            "note": "Accepted for booking.",
        },
    )
    assert accept_resp.status_code == 200
    assert accept_resp.json()["accepted_at"] is not None
    assert accept_resp.json()["accountant_reference"] == "AUG-OK"
    period_after_accept = await admin_client.get(f"/v1/admin/accounting/periods/{period_id}")
    assert period_after_accept.json()["status"] == "accepted"

    reopen_resp = await admin_client.post(
        f"/v1/admin/accounting/periods/{period_id}/reopen",
        json={"reason": "Late correction."},
    )
    assert reopen_resp.status_code == 200
    close_again_resp = await admin_client.post(f"/v1/admin/accounting/periods/{period_id}/close")
    assert close_again_resp.status_code == 200
    export_v2_resp = await admin_client.post(f"/v1/admin/accounting/periods/{period_id}/exports")
    assert export_v2_resp.status_code == 200
    assert export_v2_resp.json()["version"] == 2
    assert export_v2_resp.json()["xlsx_path"] != package["xlsx_path"]
    assert xlsx_path.exists()

    list_resp = await admin_client.get(f"/v1/admin/accounting/exports?period_id={period_id}")
    assert [item["version"] for item in list_resp.json()["items"]] == [2, 1]
    assert [item["current_final"] for item in list_resp.json()["items"]] == [True, False]

    actions = {row["action"] for row in db.execute("SELECT action FROM finance_audit_events")}
    assert "finance_export_package.generate" in actions
    assert "finance_export_package.accept" in actions
